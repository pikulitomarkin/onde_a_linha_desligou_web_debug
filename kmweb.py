from flask import Flask, render_template, request, redirect, url_for, send_from_directory
from pathlib import Path
from openpyxl import load_workbook
import folium
import gpxpy
import os
import platform

app = Flask(__name__)
app.config['STATIC_FOLDER'] = 'static'
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['EXPLAIN_TEMPLATE_LOADING'] = True
# Garante que o diretório 'static' existe
os.makedirs(app.config['STATIC_FOLDER'], exist_ok=True)
app.jinja_env.autoescape = True

class KMAppCore:
    def __init__(self):
        self.android_activity = None
        self.linhas_gpx = {
            "cmo_apa": "cmo_apucarana.gpx",
            "cmo_mga": "cmo_maringa.gpx",
            "cmo_sos": "cmo_salto_osorio.gpx",
            "cmo_sos c2": "cmo_salto_osorio_c2.gpx",
            "cmo_ssa": "cmo_salto_santiago.gpx",
            "cmo_ssac2": "cmo_salto_santiago_c2.gpx",
            "ivp_cvo": "ivp_cascavel.gpx",
            "are_ivp": "areia_ivaipora.gpx",
            "lonlns": "londrina_lns.gpx",
            "lonlna": "londrina_lna.gpx",
            "lonlna2": "londrina_lna2.gpx",
            "lon_sdi": "lon_sdi.gpx",
            "lon_mga": "lon_mga.gpx",
            "assis_c2_londrina_norte": "assis2.gpx",
            "ivp_lon": "ivaipora_londrina.gpx",
            "apucarana": "apucarana.gpx",
            "cvo_cvo": "cvo_cvo.gpx",
            "cvo_guaira": "cvo_guaira.gpx",
        }
        self.paths = None  # Adicione um atributo paths
        
    def processar_busca(self, df_key, valor_a, valor_b, col_a, col_b, nome_arquivo):
        try:
            # Validação de entrada
            try:
                km_a = float(valor_a) if valor_a else None
                km_b = float(valor_b) if valor_b else None
            except ValueError:
                raise ValueError("Os valores de KM devem ser números.")

            # Carrega a planilha
            caminho = Path(app.config['STATIC_FOLDER'], "resources") / nome_arquivo  # Ajuste aqui
            if not caminho.exists():
                raise FileNotFoundError(f"Arquivo '{nome_arquivo}' não encontrado no diretório 'resources'.")

            workbook = load_workbook(caminho, data_only=True)
            sheet = workbook.active

            # Padroniza nomes das colunas para evitar problemas com espaços/extras
            colunas = [str(cell.value).strip().upper() if cell.value else "" for cell in sheet[1]]
            col_a_proc = col_a.strip().upper()
            col_b_proc = col_b.strip().upper()

            if col_a_proc not in colunas or col_b_proc not in colunas:
                raise ValueError(f"As colunas '{col_a}' ou '{col_b}' não foram encontradas na planilha.")
            if "CODIGO" not in colunas or "MUNICIPIO" not in colunas or "SETOR" not in colunas:
                raise ValueError("As colunas obrigatórias 'CODIGO', 'MUNICIPIO' ou 'SETOR' estão ausentes.")

            idx_col_a = colunas.index(col_a_proc)
            idx_col_b = colunas.index(col_b_proc)
            idx_codigo = colunas.index("CODIGO")
            idx_municipio = colunas.index("MUNICIPIO")
            idx_setor = colunas.index("SETOR")

            menor_dif = float("inf")
            linha_selecionada = None

            # LOG para depuração: Verificar valores de entrada
            print(f"Valores fornecidos: valor_a={valor_a}, valor_b={valor_b}, col_a={col_a}, col_b={col_b}, nome_arquivo={nome_arquivo}")

            # LOG para depuração: Verificar colunas disponíveis na planilha
            print(f"Colunas disponíveis na planilha: {colunas}")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                val_a = row[idx_col_a] if km_a is not None else None
                val_b = row[idx_col_b] if km_b is not None else None
                codigo = row[idx_codigo]

                # Ignora linhas sem código de torre válido
                if not codigo or not isinstance(codigo, str):
                    continue

                # LOG para depuração
                print(f"Row: {row}")
                print(f"val_a: {val_a}, val_b: {val_b}, km_a: {km_a}, km_b: {km_b}")

                if val_a is not None and km_a is not None and abs(val_a - km_a) < menor_dif:
                    menor_dif = abs(val_a - km_a)
                    linha_selecionada = row
                if val_b is not None and km_b is not None and abs(val_b - km_b) < menor_dif:
                    menor_dif = abs(val_b - km_b)
                    linha_selecionada = row

            # LOG para depuração: Verificar linha selecionada
            if linha_selecionada:
                print(f"Linha selecionada: {linha_selecionada}")
            else:
                print("Nenhuma linha foi selecionada.")

            if linha_selecionada is None:
                raise ValueError("Nenhuma linha encontrada para os valores fornecidos.")

            # Obtem o código da torre, cidade e setor
            codigo_torre = linha_selecionada[idx_codigo]
            cidade = linha_selecionada[idx_municipio]
            setor = linha_selecionada[idx_setor]

            # LOG para depuração
            print(f"linha_selecionada: {linha_selecionada}")
            print(f"codigo_torre: {codigo_torre}, cidade: {cidade}, setor: {setor}")

            if not codigo_torre:
                raise ValueError("Código da torre não encontrado na planilha.")

            # Extrair apenas o número da torre para exibição
            numero_torre_display = self.extrair_apenas_numero(codigo_torre)
            
            # Ajustar o código da torre para o formato do GPX
            try:
                codigo_torre_ajustado = self.ajustar_codigo_torre(codigo_torre, df_key)
            except ValueError as e:
                return f"Erro ao ajustar o código da torre: {e}"

            # Buscar as coordenadas no GPX
            gpx_file = self.linhas_gpx.get(df_key)
            if not gpx_file:
                return f"Erro: Arquivo GPX não especificado para a linha {df_key}."

            torre_coords = self.buscar_torre_no_gpx(codigo_torre_ajustado, gpx_file)
            if not torre_coords:
                return render_template("resultado.html", mensagem=f"Erro: Torre '{codigo_torre_ajustado}' não encontrada no arquivo GPX.")

            latitude, longitude = torre_coords
            print("="*50)
            print("DADOS PARA O TEMPLATE detalhes_torre.html:")
            print(f"Torre original: {codigo_torre}")
            print(f"Torre para exibição: {numero_torre_display}")
            print(f"Cidade: {cidade}")
            print(f"Setor: {setor}")
            print(f"Latitude: {latitude}")
            print(f"Longitude: {longitude}")
            print(f"df_key: {df_key}")
            print("="*50)
            
            return render_template("detalhes_torre.html", detalhes={
                "Torre": numero_torre_display,  # Usar apenas o número para exibição
                "TorreOriginal": codigo_torre,  # Manter o código original para o GPX
                "Cidade": cidade,
                "Setor": setor,
                "latitude": latitude,
                "longitude": longitude
            }, df_key=df_key)

        except Exception as e:
            # Exibe o erro em um template amigável
            return render_template("resultado.html", mensagem=f"Erro: {e}")

    def mostrar_detalhes_torre(self, df_key, codigo_torre, cidade, setor):
        # Extrai apenas o número da torre
        numero_torre = self.extrair_numero_torre(codigo_torre, incluir_prefixo=False)

        # Exibe os detalhes da torre
        detalhes = {
            "Torre": numero_torre,  # Mostra apenas o número da torre
            "Cidade": cidade,
            "Setor": setor
        }

        return detalhes

    def visualizar_no_mapa(self, df_key, codigo_torre):
        try:
            # Busca a localização da torre no GPX
            gpx_file = self.linhas_gpx.get(df_key)
            if not gpx_file:
                raise ValueError("Arquivo GPX não especificado para a linha selecionada.")

            gpx_path = Path(app.config['STATIC_FOLDER'], "resources") / gpx_file  # Ajuste aqui
            if not gpx_path.exists():
                raise FileNotFoundError(f"Arquivo GPX '{gpx_file}' não encontrado no diretório 'resources'.")

            # Determina se o prefixo deve ser incluído na comparação
            incluir_prefixo = df_key in ["cmo_sos", "cmo_sos c2", "cmo_ssa", "cmo_ssac2", "lna", "apucarana", ]

            # Ajusta o código da torre para o formato esperado no GPX
            codigo_torre_ajustado = self.ajustar_codigo_torre(codigo_torre, df_key)
            if not codigo_torre_ajustado:
                raise ValueError("Código da torre ajustado é inválido.")

            # Busca a localização da torre no GPX
            torre_coords = self.buscar_torre_no_gpx(codigo_torre_ajustado, gpx_file, incluir_prefixo)
            if not torre_coords:
                raise ValueError(f"Torre '{codigo_torre}' não encontrada no arquivo GPX.")

            print(f"Coordenadas da torre encontradas: {torre_coords}")  # Log para depuração

            # Cria o mapa com folium, carrega o GPX e destaca o ponto da torre
            mapa = folium.Map(location=torre_coords, zoom_start=15)
            with open(gpx_path, "r") as gpx_file:
                gpx = gpxpy.parse(gpx_file)
                for track in gpx.tracks:
                    for segment in track.segments:
                        pontos = [(point.latitude, point.longitude) for point in segment.points]
                        folium.PolyLine(pontos, color="blue", weight=2.5, opacity=1).add_to(mapa)
            folium.Marker(
                location=torre_coords,
                popup="Torre",
                icon=folium.Icon(color="red", icon="info-sign")
            ).add_to(mapa)

            # Salva o mapa em um arquivo HTML temporário
            mapa_path = os.path.join(app.config['STATIC_FOLDER'], "mapa_torre.html")
            mapa.save(mapa_path)

            return mapa_path

        except Exception as e:
            print(f"Erro: {e}")  # Log para depuração
            return f"Erro: {e}"

    def buscar_torre_no_gpx(self, codigo_torre, gpx_file, incluir_prefixo=False):
        if not codigo_torre:
            raise ValueError("Código da torre é inválido ou não foi fornecido.")

        gpx_path = Path(app.config['STATIC_FOLDER'], "resources") / gpx_file
        if not gpx_path.exists():
            raise FileNotFoundError(f"Arquivo GPX '{gpx_file}' não encontrado no diretório 'resources'.")

        # Se já for só o número, use direto
        if codigo_torre.isdigit():
            numero_torre = codigo_torre.strip()
        else:
            numero_torre = str(self.extrair_numero_torre(codigo_torre, incluir_prefixo=False)).strip()
        print(f"Número da torre extraído da planilha (apenas número): '{numero_torre}'")

        with open(gpx_path, "r", encoding="utf-8") as gpx_file:
            gpx = gpxpy.parse(gpx_file)
            print("Waypoints disponíveis no GPX:")
            for waypoint in gpx.waypoints:
                nome_wp = str(waypoint.name).strip()
                print(f"Comparando '{nome_wp}' com '{numero_torre}'")
                if nome_wp == numero_torre:
                    print(f"Torre encontrada: {waypoint.name}")
                    return (waypoint.latitude, waypoint.longitude)

        print("Nenhuma torre correspondente encontrada no GPX.")
        return None

    def ajustar_codigo_torre(self, codigo_torre, df_key):
        """
        Ajusta o código da torre da planilha para o formato esperado no GPX.
        Exemplo:
        7350TO001 -> 001 (extrai apenas o número da torre)
        V0006R -> 6 (extrai apenas o número principal, ignorando prefixos e sufixos)
        """
        if not codigo_torre:
            raise ValueError("Código da torre é inválido ou não foi fornecido.")

        # Remove espaços em branco no início e no final
        codigo_torre = codigo_torre.strip()

        # Caso contenha "TO", extrai o número após "TO"
        if "TO" in codigo_torre:
            partes = codigo_torre.split("TO")
            if len(partes) > 1:
                numero = ''.join(filter(str.isdigit, partes[1]))  # Mantém apenas os dígitos
                if not numero:
                    raise ValueError("Número da torre não encontrado após o prefixo 'TO'.")
                return str(int(numero))  # Remove zeros à esquerda e retorna o número

        # Caso contenha apenas letras e números, extrai o número principal
        numero = ''.join(filter(str.isdigit, codigo_torre))  # Mantém apenas os dígitos
        if not numero:
            raise ValueError("Número da torre não encontrado no código fornecido.")

        return str(int(numero))  # Remove zeros à esquerda e retorna o número

    def extrair_numero_torre(self, codigo_torre, incluir_prefixo=False):
        """
        Extrai o número da torre, mantendo letras sufixo, se presentes.
        Exemplo:
        V0006R -> 0006R
        """
        if not codigo_torre:
            return None

        codigo_torre = codigo_torre.strip()

        # Caso contenha "TO", extrai o número após "TO"
        if "TO" in codigo_torre:
            partes = codigo_torre.split("TO")
            if len(partes) > 1:
                prefixo = partes[0] if incluir_prefixo else ""
                numero = ''.join(filter(str.isdigit, partes[1]))  # Mantém apenas os dígitos
                if not numero:
                    return None
                numero = str(int(numero))  # Remove zeros à esquerda
                return f"{prefixo}{numero}"

        # Caso não contenha "TO", tenta extrair números e manter sufixos alfanuméricos
        numero = ''.join(filter(str.isalnum, codigo_torre))  # Mantém apenas caracteres alfanuméricos
        if not numero:
            return None

        return numero

    def extrair_apenas_numero(self, codigo_torre):
        """
        Extrai apenas o número da torre, removendo prefixos mas mantendo sufixos.
        Exemplos:
        V0005 -> 5
        7460TO006 -> 6
        V0006R -> 6R
        """
        if not codigo_torre:
            return None

        codigo_torre = str(codigo_torre).strip()

        # Caso contenha "TO", extrai o número após "TO"
        if "TO" in codigo_torre:
            partes = codigo_torre.split("TO")
            if len(partes) > 1:
                numero = ''.join(filter(str.isdigit, partes[1]))  # Mantém apenas os dígitos
                if not numero:
                    return None
                return str(int(numero))  # Remove zeros à esquerda e retorna o número

        # Para outros casos, extrai o número principal e mantém sufixos alfabéticos
        import re
        # Padrão: captura prefixos opcionais + número + sufixos opcionais
        match = re.match(r'^[A-Z]*(\d+)([A-Z]*)$', codigo_torre.upper())
        
        if match:
            numero = match.group(1)  # Parte numérica
            sufixo = match.group(2)  # Parte alfabética após o número
            
            if numero:
                numero_limpo = str(int(numero))  # Remove zeros à esquerda
                return f"{numero_limpo}{sufixo}" if sufixo else numero_limpo
        
        # Fallback: extrai apenas os dígitos se o padrão não funcionar
        numero = ''.join(filter(str.isdigit, codigo_torre))
        if not numero:
            return None

        return str(int(numero))  # Remove zeros à esquerda e retorna o número

km_app = KMAppCore()

@app.route("/")
def index():
    return render_template("menu_principal.html")

@app.route("/londrina")
def londrina():
    botoes_londrina = [
        ("Linha Londrina - Londrina Sul", "lonlns", "KMLON", "KMAPA", "KM LON LNS.xlsx"),
        ("Linha Londrina - Londrina Norte C1", "lonlna", "KM - LON - LNA", "KM - LNA - LON", "KM LON LNA.xlsx"),
        ("Linha Londrina - Londrina Norte C2", "lonlna2", "LON-LNA", "LNA-LON", "KM LON LNA2.xlsx"),
        ("Linha Londrina Sul - Apucarana", "apucarana", "LNS", "APA", "KM LON APA.xlsx"),
        ("Linha Londrina - Sarandi", "lon_sdi", "LON-SDI", "SDI-LON", "KM LON SDI.xlsx"),
        ("Linha Maringa - Sarandi", "lon_mga", "MGA-SDI", "SDI-MGA", "KM MGA SDI.xlsx"),
        ("Linha Assis C2 - Londrina Norte", "assis_c2_londrina_norte", "LNA-ASS", "ASS-LNA", "KM LNA ASS2.xlsx"),
        ("Linha Assis C1 - Londrina Norte", "lna_assis", "KM - LNA - ASS", "KM - ASS - LNA", "KM LNA ASS.xlsx"),
        ("Linha Ivaiporã - Londrina", "ivp_lon", "KMIVP", "KMLON", "KM IVP LON.xlsx")
    ]
    return render_template("menu_londrina.html", botoes_londrina=botoes_londrina)

@app.route("/campomourao")
def campomourao():
    botoes_campomourao = [
        ("Linha Campo Mourão - Apucarana", "cmo_apa", "KMCMO", "KMAPA", "KM CMO APA.xlsx"),
        ("Linha Campo Mourão - Maringá", "cmo_mga", "KMCMO", "KMMGA", "KM CMO MGA.xlsx"),
        ("Linha Salto Osório - Campo Mourão", "cmo_sos", "KMSOS", "KMCMO", "KM CMO SOS.xlsx"),
        ("Linha Salto Osório C2 - Campo Mourão", "cmo_sos c2", "KMSOS", "KMCMO", "KM CMO SOSC2.xlsx"),
        ("Linha Salto Santiago - Campo Mourão", "cmo_ssa", "KMSSA", "KMIVP", "KM CMO SSA.xlsx"),
        ("Linha Salto Santiago C2 - Campo Mourão", "cmo_ssac2", "KMSSA", "KMIVP", "KM CMO SSAC2.xlsx"),
        ("Linha Ivaiporã - Cascavel", "ivp_cvo", "KMIVP", "KMCVO", "KM IVP CVO.xlsx"),
        ("Linha Cascavel - Cascavel Oeste", "cvo_cvo", "CEL-CVO", "CVO-CEL", "KM CEL CVO.xlsx"),
        ("Linha Cascavel - Guaira", "cvo_guaira", "CVO-GUI", "GUI-CVO", "KM CVO GUI.xlsx"),
        ("Linha Areia - Ivaiporã", "are_ivp", "KMARE", "KMIVP", "KM ARE IVP.xlsx"),
    ]
    return render_template("menu_campomourao.html", botoes_campomourao=botoes_campomourao)

@app.route("/abrir_km", methods=["POST"])
def abrir_km():
    chave = request.form["chave"]
    col_a = request.form["col_a"]
    col_b = request.form["col_b"]
    texto_linha = request.form["texto_linha"]
    nome_arquivo = request.form["nome_arquivo"]
    return render_template("abrir_km.html", chave=chave, col_a=col_a, col_b=col_b, texto_linha=texto_linha, nome_arquivo=nome_arquivo)

@app.route("/processar_busca", methods=["POST"])
def processar_busca():
    df_key = request.form["chave"]
    valor_a = request.form["valor_a"]
    valor_b = request.form["valor_b"]
    col_a = request.form["col_a"]
    col_b = request.form["col_b"]
    nome_arquivo = request.form["nome_arquivo"]

    # A função processar_busca da classe já retorna o template renderizado
    return km_app.processar_busca(df_key, valor_a, valor_b, col_a, col_b, nome_arquivo)

@app.route("/visualizar_mapa", methods=["POST"])
def visualizar_mapa():
    df_key = request.form["df_key"]
    codigo_torre = request.form["codigo_torre"]

    mapa_path = km_app.visualizar_no_mapa(df_key, codigo_torre)

    if "Erro" in mapa_path:
         return render_template("resultado.html", mensagem=mapa_path)
    else:
        # Corrige o caminho para URL estática
        mapa_url = url_for('static', filename="mapa_torre.html")
        return render_template("mapa.html", mapa_path=mapa_url)

# Rota para servir arquivos estáticos (CSS, JS, imagens, etc.)
@app.route('/static/<path:filename>')
def serve_static(filename):
    return send_from_directory(app.config['STATIC_FOLDER'], filename)

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5000))
    resources_dir = os.path.join(app.config['STATIC_FOLDER'], 'resources')
    os.makedirs(resources_dir, exist_ok=True)
    app.run(host="0.0.0.0", port=port)
